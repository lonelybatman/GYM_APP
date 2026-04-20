"""
Exercise Migration Script  v2
exercise1.csv + muscle.xlsx → Supabase

Tabellen-Logik:
  muscles      — aus muscle.xlsx, einmalig befüllt
  exercises    — eine Zeile pro Übungs-Block (exercise kombi + muscle + name)
  equipment    — eine Zeile pro Equipment-Zeile, mit JSONB details

Setup:
    pip install pandas openpyxl supabase python-dotenv

.env:
    SUPABASE_URL=https://xxxx.supabase.co
    SUPABASE_SERVICE_KEY=eyJ...
"""

import os
import json
import uuid
import math
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

EXERCISE_CSV  = "exercise1.csv"
MUSCLE_XLSX   = "muscle.xlsx"

SUPABASE_URL  = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY  = os.environ.get("SUPABASE_SERVICE_KEY", "")

# ─────────────────────────────────────────────────────────────────────────────
# SQL SCHEMA  (einmalig im Supabase SQL Editor ausführen)
# ─────────────────────────────────────────────────────────────────────────────
SCHEMA_SQL = """
-- 1) Muskeln
CREATE TABLE IF NOT EXISTS muscles (
    id           UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    name_en      TEXT NOT NULL,          -- "Biceps", "Triceps" …
    name_de      TEXT,
    name_latin   TEXT,
    created_at   TIMESTAMPTZ DEFAULT NOW()
);

-- 2) Übungen
CREATE TABLE IF NOT EXISTS exercises (
    id                UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    full_name         TEXT NOT NULL,     -- "Free + Cable > triceps > pushdown"
    exercise_kombi    TEXT,
    target_muscle     TEXT,
    exercise_name     TEXT,
    muscle_id         UUID REFERENCES muscles(id),
    is_superdefault   BOOLEAN DEFAULT FALSE,
    created_at        TIMESTAMPTZ DEFAULT NOW()
);

-- 3) Equipment-Zeilen mit JSONB
CREATE TABLE IF NOT EXISTS equipment (
    id                UUID PRIMARY KEY DEFAULT gen_random_uuid(),
    exercise_id       UUID REFERENCES exercises(id) ON DELETE CASCADE,
    equipment_name    TEXT NOT NULL,
    possible          BOOLEAN DEFAULT FALSE,
    is_default_setup  BOOLEAN DEFAULT FALSE,
    details           JSONB DEFAULT '{}',
    created_at        TIMESTAMPTZ DEFAULT NOW()
);
"""

# ─────────────────────────────────────────────────────────────────────────────
# Spalten-Index → semantischer Name  (für JSONB details, ab Spalte H = idx 7)
# ─────────────────────────────────────────────────────────────────────────────
DETAIL_COLS = {
    7:  "hands_1",
    8:  "hands_2",
    9:  "grip_sup",
    10: "grip_s_n",
    11: "grip_n",
    12: "grip_n_p",
    13: "grip_pro",
    14: "plane_lat",
    15: "plane_l_s",
    16: "plane_sag",
    17: "plane_s_t",
    18: "plane_trans",
    19: "plane_t_l",
    20: "width_x05",
    21: "width_x1",
    22: "width_x15",
    23: "cable_h_1_8",
    24: "cable_h_9_15",
    25: "cable_h_16_22",
    26: "bench_flat",
    27: "bench_incl",
    28: "bench_upright",
    29: "stand_0",
    30: "stand_45",
    31: "stand_90",
    32: "stand_90os",
    33: "stand_135",
    34: "stand_180",
    35: "bench_cable_0",
    36: "bench_cable_90",
    37: "bench_cable_180",
    38: "body_bench_0",
    39: "body_bench_45",
    40: "body_bench_90",
    41: "body_bench_135",
    42: "body_bench_180",
    43: "body_pos_lying",
    44: "body_pos_180",
    45: "body_pos_sitting",
    46: "sit_0",
    47: "sit_90",
    # 48 = info_equipment  → eigenes Feld im Equipment-Record
    # 49 = info_exercise   → wird ignoriert (gehört zur Übung, nicht Equipment)
    50: "cables_1",
    51: "cables_2",
}

# ─────────────────────────────────────────────────────────────────────────────
# Muskel-Mapping: CSV-Wert → muscle.xlsx Spalte A  (normalisiert, lowercase)
# ─────────────────────────────────────────────────────────────────────────────
MUSCLE_ALIAS = {
    "abs":      "core",        # abs ist Teil von core
    "lat":      "back",        # lat ist Teil von back
    "shoulder": "shoulders",   # CSV schreibt ohne s
}

# ─────────────────────────────────────────────────────────────────────────────
# Hilfsfunktionen
# ─────────────────────────────────────────────────────────────────────────────

def is_true(val) -> bool:
    """WAHR / WAHR* / True / 1  →  True;  alles andere → False"""
    if val is True:
        return True
    if isinstance(val, str):
        return val.strip() in ("WAHR", "WAHR*")
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return val == 1
    return False

def is_wahr_star(val) -> bool:
    """Nur exakt WAHR* → True  (für is_default_setup)"""
    return isinstance(val, str) and val.strip() == "WAHR*"

def clean(val) -> str | None:
    """Trimmt Strings, gibt None zurück wenn leer/nan"""
    if val is None:
        return None
    if isinstance(val, float) and math.isnan(val):
        return None
    s = str(val).strip()
    return s if s and s.lower() != "nan" else None

def is_empty_row(row) -> bool:
    return all(
        clean(row[i]) in (None, "FALSCH", "FALSE")
        for i in range(52)
    )

def is_sub_header(row) -> bool:
    """Sub-Header-Zeile: Spalte H = '1', Spalte I = '2'"""
    h = clean(row[7])
    i = clean(row[8])
    return h in ("1", "1.0") and i in ("2", "2.0")

# ─────────────────────────────────────────────────────────────────────────────
# muscle.xlsx lesen
# ─────────────────────────────────────────────────────────────────────────────

def parse_muscles(path: str) -> tuple[list[dict], dict[str, str]]:
    """
    Gibt (muscle_records, name_to_id) zurück.
    name_to_id: {"biceps": "<uuid>", "triceps": "<uuid>", ...}
    """
    wb = load_workbook(path, read_only=True)
    ws = wb.active

    records = []
    name_to_id: dict[str, str] = {}
    current_name_en = None

    for i, row in enumerate(ws.iter_rows(max_row=50, values_only=True)):
        if i < 2:
            continue  # Header
        if row[0]:
            current_name_en = str(row[0]).strip()

        if not current_name_en:
            continue

        mid = str(uuid.uuid4())
        records.append({
            "id":          mid,
            "name_en":     current_name_en,
            "name_de":     clean(row[1]),
            "name_latin":  clean(row[2]),
        })

        key = current_name_en.strip().lower()
        if key not in name_to_id:
            name_to_id[key] = mid

    return records, name_to_id

# ─────────────────────────────────────────────────────────────────────────────
# exercise1.csv lesen
# ─────────────────────────────────────────────────────────────────────────────

def parse_exercises(path: str, name_to_id: dict[str, str]) -> tuple[list[dict], list[dict]]:
    """
    Gibt (exercise_records, equipment_records) zurück.
    """
    df = pd.read_csv(path, sep=";", header=None, encoding="utf-8-sig", dtype=str)

    exercise_records: list[dict] = []
    equipment_records: list[dict] = []

    current_exercise_id: str | None = None
    current_kombi: str | None = None

    for i in range(4, len(df)):
        row = list(df.iloc[i])

        if is_sub_header(row) or is_empty_row(row):
            continue

        c_kombi   = clean(row[0])   # A — exercise kombi
        c_super   = clean(row[1])   # B — Superdefault
        c_muscle  = clean(row[2])   # C — Target Muscle
        c_name    = clean(row[3])   # D — exercise name
        c_poss    = clean(row[4])   # E — possible
        c_equip   = clean(row[5])   # F — equipment name
        c_default = clean(row[6])   # G — default

        # Neuer exercise-Block wenn Spalte C (muscle) gesetzt ist
        if c_muscle and c_muscle.lower() not in ("target muscle",):
            if c_kombi:
                current_kombi = c_kombi

            muscle_key = c_muscle.strip().lower()
            muscle_key = MUSCLE_ALIAS.get(muscle_key, muscle_key)
            muscle_id  = name_to_id.get(muscle_key)

            if muscle_id is None:
                print(f"  ⚠ Muskel nicht gefunden: '{c_muscle}' (Zeile {i+1})")

            full_name = " > ".join(filter(None, [
                current_kombi,
                c_muscle.strip() if c_muscle else None,
                c_name.strip()   if c_name   else None,
            ]))

            current_exercise_id = str(uuid.uuid4())
            exercise_records.append({
                "id":               current_exercise_id,
                "full_name":        full_name,
                "exercise_kombi":   current_kombi,
                "target_muscle":    c_muscle.strip() if c_muscle else None,
                "exercise_name":    c_name.strip()   if c_name   else None,
                "muscle_id":        muscle_id,
                "is_superdefault":  is_true(c_super),
            })

            # Equipment auf derselben Zeile?
            if c_equip:
                equipment_records.append(
                    _build_equipment(row, current_exercise_id, c_equip, c_poss, c_default)
                )

        # Weitere Equipment-Zeilen (keine neue Übung)
        elif c_equip and current_exercise_id:
            equipment_records.append(
                _build_equipment(row, current_exercise_id, c_equip, c_poss, c_default)
            )

    return exercise_records, equipment_records


def _build_equipment(row, exercise_id: str, equip_name: str,
                     possible_val, default_val) -> dict:
    # JSONB details: alle Boolean-Felder ab Spalte H
    details: dict[str, bool] = {}
    for col_idx, field_name in DETAIL_COLS.items():
        if col_idx < len(row):
            details[field_name] = is_true(row[col_idx])

    # info_equipment als eigenes Textfeld im details-Objekt
    info = clean(row[48]) if len(row) > 48 else None
    if info:
        details["info_equipment"] = info

    return {
        "id":               str(uuid.uuid4()),
        "exercise_id":      exercise_id,
        "equipment_name":   equip_name.strip(),
        "possible":         is_true(possible_val),
        "is_default_setup": is_wahr_star(default_val),
        "details":          details,
    }

# ─────────────────────────────────────────────────────────────────────────────
# Supabase Upload
# ─────────────────────────────────────────────────────────────────────────────

def upload(muscles, exercises, equipment, dry_run: bool = False):
    print(f"\n{'[DRY RUN] ' if dry_run else ''}Zusammenfassung:")
    print(f"  {len(muscles):4d}  muscles")
    print(f"  {len(exercises):4d}  exercises")
    print(f"  {len(equipment):4d}  equipment-Einträge")

    if dry_run:
        print("\n── Beispiel exercise ──")
        print(json.dumps(exercises[0], indent=2, ensure_ascii=False))
        print("\n── Beispiel equipment ──")
        if equipment:
            sample = dict(equipment[0])
            sample["details"] = {k: v for k, v in list(sample["details"].items())[:5]}
            sample["details"]["..."] = "..."
            print(json.dumps(sample, indent=2, ensure_ascii=False))
        print("\nDry run — kein Upload.")
        return

    from supabase import create_client
    client = create_client(SUPABASE_URL, SUPABASE_KEY)

    BATCH = 500

    print("\nLade muscles …")
    for k in range(0, len(muscles), BATCH):
        client.table("muscles").upsert(muscles[k:k+BATCH]).execute()
    print(f"  ✓ {len(muscles)} muscles")

    print("Lade exercises …")
    for k in range(0, len(exercises), BATCH):
        client.table("exercises").upsert(exercises[k:k+BATCH]).execute()
    print(f"  ✓ {len(exercises)} exercises")

    print("Lade equipment …")
    for k in range(0, len(equipment), BATCH):
        # details muss als JSON-String übergeben werden
        batch = []
        for eq in equipment[k:k+BATCH]:
            e = dict(eq)
            e["details"] = json.dumps(e["details"], ensure_ascii=False)
            batch.append(e)
        client.table("equipment").upsert(batch).execute()
    print(f"  ✓ {len(equipment)} equipment")

    print("\n✅ Migration abgeschlossen!")

# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Exercise Migration v2")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parsen & prüfen, kein Upload")
    parser.add_argument("--schema",  action="store_true",
                        help="SQL Schema ausgeben und beenden")
    args = parser.parse_args()

    if args.schema:
        print(SCHEMA_SQL)
        raise SystemExit(0)

    print("📋 Lese muscle.xlsx …")
    muscles, name_to_id = parse_muscles(MUSCLE_XLSX)
    print(f"   {len(muscles)} Muskel-Einträge, {len(name_to_id)} eindeutige Gruppen")

    print("📂 Lese exercise1.csv …")
    exercises, equipment = parse_exercises(EXERCISE_CSV, name_to_id)
    print(f"   {len(exercises)} Übungen, {len(equipment)} Equipment-Zeilen")

    upload(muscles, exercises, equipment, dry_run=args.dry_run)
